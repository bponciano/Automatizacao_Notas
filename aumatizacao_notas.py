# Automatizador de Notas

#Pacotes
import os
import shutil
import pandas as pd
import xml.etree.ElementTree as ET
import sqlite3
from openpyxl import load_workbook

#Caminhos
entrada = '/workspaces/automatizacao_nota/dados_brutos'
processados = '/workspaces/automatizacao_nota/dados_processados'
data_base = '/workspaces/automatizacao_nota/Metalica_DataBase.db'
planilha = '/workspaces/automatizacao_nota/notas.xlsx'

os.makedirs(entrada, exist_ok=True)
os.makedirs(processados, exist_ok=True)

print("Ambiente configurado com sucesso!")

#Banco de Dados e Tabelas

def processar_xml(caminho_xml, cursor):

    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
    tree = ET.parse(caminho_xml)
    root = tree.getroot()

    if root.find('.//nfe:infNFe', ns) is None:
        return []

    lista_para_excel = []

    emit = root.find('.//nfe:emit', ns)
    cnpj = emit.find('nfe:CNPJ', ns).text
    razao = emit.find('nfe:xNome', ns).text
    
    end = emit.find('nfe:enderEmit', ns)
    rua, nro, bairro, cid, uf = (
        end.find('nfe:xLgr', ns).text, end.find('nfe:nro', ns).text, 
        end.find('nfe:xBairro', ns).text, end.find('nfe:xMun', ns).text, 
        end.find('nfe:UF', ns).text
    )

    #Inserindo o Endereço
    cursor.execute('INSERT OR IGNORE INTO Enderecos (Rua, Numero, Bairro, Cidade, Estado) VALUES (?,?,?,?,?)', (rua, nro, bairro, cid, uf))
    cursor.execute('SELECT ID_Endereco FROM Enderecos WHERE Rua=? AND Numero=? AND Bairro=? AND Cidade=?', (rua, nro, bairro, cid))
    id_end = cursor.fetchone()[0]

    #Inserido o Fornecedor
    cursor.execute('INSERT OR IGNORE INTO Fornecedores (CNPJ, Razao_Social, ID_Endereco) VALUES (?,?,?)', (cnpj, razao, id_end))
    cursor.execute('SELECT ID_Fornecedor FROM Fornecedores WHERE CNPJ=?', (cnpj,))
    id_forn = cursor.fetchone()[0]

    #Extraindo dados da Nota
    ide = root.find('.//nfe:ide', ns)
    n_nota = ide.find('nfe:nNF', ns).text
    dt_emi = ide.find('nfe:dhEmi', ns).text[:10]
    v_nota = float(root.find('.//nfe:total/nfe:ICMSTot/nfe:vNF', ns).text)

    #Inserindo dados da nota
    primeiro_det = root.find('.//nfe:det', ns)
    cfop_base = primeiro_det.find('.//nfe:CFOP', ns).text if primeiro_det is not None else ""

    if cfop_base.startswith(('11', '21', '51', '61')):
        tipo_nota = "VENDA"
    elif cfop_base.startswith(('19', '29', '59', '69')):
        tipo_nota = "REMESSA"
    else:
        tipo_nota = "OUTROS"

    cursor.execute('''INSERT OR IGNORE INTO Notas (Num_Nota, Data, Valor, ID_Fornecedor, Tipo_Nota) 
                      VALUES (?,?,?,?,?)''', 
                   (n_nota, dt_emi, v_nota, id_forn, tipo_nota))
    
    cursor.execute('SELECT ID_Nota FROM Notas WHERE Num_Nota=? AND ID_Fornecedor=?', (n_nota, id_forn))
    id_nota = cursor.fetchone()[0]

    #Extraindo itens da nota
    for det in root.findall('.//nfe:det', ns):
        prod = det.find('nfe:prod', ns)
        nome_p = prod.find('nfe:xProd', ns).text
        qtd = float(prod.find('nfe:qCom', ns).text)
        v_un = float(prod.find('nfe:vUnCom', ns).text)
        v_tot = float(prod.find('nfe:vProd', ns).text) 

        cursor.execute('INSERT OR IGNORE INTO Produtos (Nome) VALUES (?)', (nome_p,))
        cursor.execute('SELECT ID_Produto FROM Produtos WHERE Nome=?', (nome_p,))
        id_prod = cursor.fetchone()[0]

        cursor.execute('''INSERT INTO Compras (ID_Nota, ID_Fornecedor, ID_Produto, Data, Quantidade, Valor_Unitario, Valor_Total)
                          VALUES (?,?,?,?,?,?,?)''', 
                       (id_nota, id_forn, id_prod, dt_emi, qtd, v_un, v_tot))

        lista_para_excel.append([cnpj, razao, dt_emi, n_nota, nome_p, qtd, v_un, v_tot, tipo_nota])

    return lista_para_excel

#Atualização Excel

conn = sqlite3.connect(data_base)
cursor = conn.cursor()

arquivos = [f for f in os.listdir(entrada) if f.lower().endswith('.xml')]
lista_consolidada = []

for arquivo in arquivos:
    caminho_completo = os.path.join(entrada, arquivo)
    print(f"Processando: {arquivo}")

    try:
        dados_da_nota = processar_xml(caminho_completo, cursor)
        lista_consolidada.extend(dados_da_nota)
        conn.commit()
        shutil.move(caminho_completo, os.path.join(processados, arquivo))

    except Exception as e:
        conn.rollback()
        print(f"Erro ao processar {arquivo}: {e}")

conn.close()

if lista_consolidada:
    colunas = ['CNPJ', 'Fornecedor', 'Data', 'Nº Nota', 'Produto', 'Qtd', 'V.Unit', 'V.Total', 'Tipo']
    df_novos = pd.DataFrame(lista_consolidada, columns=colunas)

    if not os.path.exists(planilha):
        df_novos.to_excel(planilha, index=False, sheet_name='Notas')

        wb = load_workbook(planilha)
        ws = wb['Notas']
        from openpyxl.worksheet.table import Table, TableStyleInfo

        ref = f"A1:{chr(64 + len(colunas))}{len(df_novos) + 1}"
        tab = Table(displayName="TabelaNotas", ref=ref)

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(planilha)

    else:
        wb = load_workbook(planilha)
        ws = wb.active

        start_row = ws.max_row + 1

        for r_idx, row in enumerate(df_novos.values, start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        if ws.tables:
            nome_tabela = list(ws.tables.keys())[0]
            tabela = ws.tables[nome_tabela]
            nova_ref = f"A1:{chr(64 + len(colunas))}{ws.max_row}"
            tabela.ref = nova_ref

        wb.save(planilha)
            
    print(f"Sucesso! {len(lista_consolidada)} itens adicionados ao Excel e SQL com formatação mantida.")

else:
    print("Nenhum novo XML para processar.")